"""
Tests for Excel Processor
Run with: pytest test_excel_processor.py
"""

import pytest
import openpyxl
from excel_processor import ExcelProcessor, ActionPlanner
import os
import tempfile

@pytest.fixture
def sample_workbook():
    """Create a sample Excel workbook for testing"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TestData"
    
    # Add headers
    ws.append(["Name", "Email", "Phone", "Amount"])
    
    # Add sample data with issues
    ws.append(["  John Smith  ", "john@example.com", "1234567890", "100.50"])
    ws.append(["Jane Doe", "jane@test.com  ", "0987654321", "200.75"])
    ws.append(["  John Smith  ", "john@example.com", "1234567890", "100.50"])  # Duplicate
    ws.append(["Bob Wilson", "bob@company.org", "(123) 456-7890", "150.25"])
    
    # Save to temp file
    temp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(temp.name)
    temp.close()
    
    yield temp.name
    
    # Cleanup
    if os.path.exists(temp.name):
        os.remove(temp.name)


class TestExcelProcessor:
    """Test Excel Processor functionality"""
    
    def test_processor_initialization(self, sample_workbook):
        """Test that processor can load a workbook"""
        processor = ExcelProcessor(sample_workbook)
        assert processor.workbook is not None
        assert len(processor.workbook.sheetnames) > 0
    
    def test_trim_clean(self, sample_workbook):
        """Test text cleaning functionality"""
        processor = ExcelProcessor(sample_workbook)
        
        # Execute trim/clean
        processor._trim_clean({"sheet": "TestData"})
        
        # Check that spaces were removed
        sheet = processor.workbook["TestData"]
        name_cell = sheet.cell(row=2, column=1).value
        
        assert name_cell == "John Smith"  # No extra spaces
        assert "Cleaned text in sheet" in processor.changes_log[0]
    
    def test_remove_duplicates(self, sample_workbook):
        """Test duplicate removal"""
        processor = ExcelProcessor(sample_workbook)
        
        # Get initial row count
        initial_rows = processor.workbook["TestData"].max_row
        
        # Remove duplicates
        processor._trim_clean({"sheet": "TestData"})
        processor._remove_duplicates({"sheet": "TestData"})
        
        # Check that duplicate was removed
        final_rows = processor.workbook["TestData"].max_row
        assert final_rows < initial_rows
    
    def test_execute_plan(self, sample_workbook):
        """Test executing a complete plan"""
        processor = ExcelProcessor(sample_workbook)
        
        plan = [
            {
                "type": "trim_clean",
                "description": "Clean text",
                "params": {}
            },
            {
                "type": "remove_duplicates",
                "description": "Remove dupes",
                "params": {}
            }
        ]
        
        result = processor.execute_plan(plan)
        
        assert result["success"] is True
        assert result["actions_completed"] == 2
        assert len(result["changes"]) == 2
    
    def test_get_diff_summary(self, sample_workbook):
        """Test diff summary generation"""
        processor = ExcelProcessor(sample_workbook)
        
        processor._trim_clean({})
        diff = processor.get_diff_summary()
        
        assert "changes" in diff
        assert len(diff["changes"]) > 0


class TestActionPlanner:
    """Test Action Planner functionality"""
    
    def test_parse_duplicate_removal(self):
        """Test parsing duplicate removal request"""
        plan = ActionPlanner.parse_request("Remove duplicates")
        
        assert len(plan) > 0
        assert any(action["type"] == "remove_duplicates" for action in plan)
    
    def test_parse_clean_request(self):
        """Test parsing cleaning request"""
        plan = ActionPlanner.parse_request("Clean and trim the data")
        
        assert any(action["type"] == "trim_clean" for action in plan)
    
    def test_parse_pivot_request(self):
        """Test parsing pivot table request"""
        plan = ActionPlanner.parse_request("Create pivot table")
        
        assert any(action["type"] == "create_pivot" for action in plan)
    
    def test_parse_complex_request(self):
        """Test parsing complex multi-operation request"""
        plan = ActionPlanner.parse_request(
            "Remove duplicates, clean data, and create pivot table"
        )
        
        assert len(plan) >= 2
        action_types = [action["type"] for action in plan]
        assert "remove_duplicates" in action_types or "trim_clean" in action_types


if __name__ == "__main__":
    pytest.main([__file__, "-v"])

