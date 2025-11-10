"""
Excel Processing Engine
Handles all Excel file manipulations using openpyxl and pandas
"""

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
import pandas as pd
from typing import List, Dict, Any, Optional
import re
from datetime import datetime


class ExcelProcessor:
    """Main Excel processing engine"""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(file_path)
        self.changes_log = []
    
    def execute_plan(self, plan: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Execute a series of Excel actions based on the plan"""
        results = {
            "success": True,
            "actions_completed": 0,
            "changes": [],
            "errors": []
        }
        
        for action in plan:
            try:
                action_type = action.get("type")
                params = action.get("params", {})
                
                if action_type == "trim_clean":
                    self._trim_clean(params)
                elif action_type == "remove_duplicates":
                    self._remove_duplicates(params)
                elif action_type == "split_column":
                    self._split_column(params)
                elif action_type == "create_pivot":
                    self._create_pivot(params)
                elif action_type == "standardize_phone":
                    self._standardize_phone(params)
                elif action_type == "convert_dates":
                    self._convert_dates(params)
                elif action_type == "add_calculated_column":
                    self._add_calculated_column(params)
                else:
                    results["errors"].append(f"Unknown action type: {action_type}")
                    continue
                
                results["actions_completed"] += 1
                results["changes"].append(action.get("description", action_type))
                
            except Exception as e:
                results["errors"].append(f"Error in {action_type}: {str(e)}")
                results["success"] = False
        
        return results
    
    def _trim_clean(self, params: Dict[str, Any]):
        """Remove leading/trailing spaces and clean non-printable characters"""
        sheet_name = params.get("sheet") or self.workbook.sheetnames[0]
        sheet = self.workbook[sheet_name]
        
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    # Trim and clean
                    cleaned = cell.value.strip()
                    cleaned = re.sub(r'[\x00-\x1f\x7f-\x9f]', '', cleaned)
                    cell.value = cleaned
        
        self.changes_log.append(f"Cleaned text in sheet: {sheet_name}")
    
    def _remove_duplicates(self, params: Dict[str, Any]):
        """Remove duplicate rows"""
        sheet_name = params.get("sheet") or self.workbook.sheetnames[0]
        sheet = self.workbook[sheet_name]
        
        # Convert to pandas for easier duplicate handling
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
        
        df = pd.DataFrame(data[1:], columns=data[0])
        original_count = len(df)
        
        # Remove duplicates
        df = df.drop_duplicates(keep='first')
        removed_count = original_count - len(df)
        
        # Clear sheet and write back
        sheet.delete_rows(2, sheet.max_row)
        for r_idx, row in enumerate(df.values, start=2):
            for c_idx, value in enumerate(row, start=1):
                sheet.cell(row=r_idx, column=c_idx, value=value)
        
        self.changes_log.append(f"Removed {removed_count} duplicate rows from {sheet_name}")
    
    def _split_column(self, params: Dict[str, Any]):
        """Split a column into multiple columns"""
        sheet_name = params.get("sheet") or self.workbook.sheetnames[0]
        source_col = params.get("source_col")
        into = params.get("into", ["Part1", "Part2"])
        delimiter = params.get("delimiter", " ")
        
        sheet = self.workbook[sheet_name]
        
        # Find source column
        header_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        try:
            col_idx = header_row.index(source_col) + 1
        except ValueError:
            raise ValueError(f"Column '{source_col}' not found")
        
        # Add new column headers
        last_col = sheet.max_column
        for i, new_col_name in enumerate(into):
            sheet.cell(row=1, column=last_col + i + 1, value=new_col_name)
        
        # Split data
        for row_idx in range(2, sheet.max_row + 1):
            cell_value = sheet.cell(row=row_idx, column=col_idx).value
            if cell_value and isinstance(cell_value, str):
                parts = cell_value.split(delimiter)
                for i, part in enumerate(parts[:len(into)]):
                    sheet.cell(row=row_idx, column=last_col + i + 1, value=part.strip())
        
        self.changes_log.append(f"Split column '{source_col}' into {len(into)} columns")
    
    def _create_pivot(self, params: Dict[str, Any]):
        """Create a pivot table summary (simplified version)"""
        sheet_name = params.get("sheet") or self.workbook.sheetnames[0]
        sheet = self.workbook[sheet_name]
        
        # Convert to pandas for pivot
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
        
        df = pd.DataFrame(data[1:], columns=data[0])
        
        # Create pivot
        rows = params.get("rows", [])
        values = params.get("values", [])
        
        if rows and values:
            pivot = df.pivot_table(
                index=rows,
                values=[v["field"] for v in values],
                aggfunc={v["field"]: v.get("agg", "sum").lower() for v in values}
            )
            
            # Create new sheet for pivot
            pivot_sheet_name = params.get("destination", "Pivot_Summary")
            if pivot_sheet_name in self.workbook.sheetnames:
                del self.workbook[pivot_sheet_name]
            
            pivot_sheet = self.workbook.create_sheet(pivot_sheet_name)
            
            # Write pivot to new sheet
            # Headers
            pivot_sheet.cell(row=1, column=1, value=rows[0])
            for i, col in enumerate(pivot.columns):
                pivot_sheet.cell(row=1, column=i + 2, value=str(col))
            
            # Data
            for r_idx, (idx, row) in enumerate(pivot.iterrows(), start=2):
                pivot_sheet.cell(row=r_idx, column=1, value=str(idx))
                for c_idx, value in enumerate(row.values, start=2):
                    pivot_sheet.cell(row=r_idx, column=c_idx, value=value)
            
            self.changes_log.append(f"Created pivot table in sheet: {pivot_sheet_name}")
    
    def _standardize_phone(self, params: Dict[str, Any]):
        """Standardize phone number format"""
        sheet_name = params.get("sheet") or self.workbook.sheetnames[0]
        phone_col = params.get("phone_col", "Phone")
        country_code = params.get("country_code", "234")
        
        sheet = self.workbook[sheet_name]
        
        # Find phone column
        header_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        try:
            col_idx = header_row.index(phone_col) + 1
        except ValueError:
            raise ValueError(f"Column '{phone_col}' not found")
        
        # Standardize format
        for row_idx in range(2, sheet.max_row + 1):
            cell_value = sheet.cell(row=row_idx, column=col_idx).value
            if cell_value:
                # Remove all non-digits
                digits = re.sub(r'\D', '', str(cell_value))
                
                # Add country code if missing
                if not digits.startswith(country_code):
                    digits = country_code + digits[-10:]
                
                # Format: +234-XXX-XXX-XXXX
                if len(digits) >= 10:
                    formatted = f"+{digits[:3]}-{digits[3:6]}-{digits[6:9]}-{digits[9:]}"
                    sheet.cell(row=row_idx, column=col_idx, value=formatted)
        
        self.changes_log.append(f"Standardized phone numbers in column: {phone_col}")
    
    def _convert_dates(self, params: Dict[str, Any]):
        """Convert and standardize date formats"""
        sheet_name = params.get("sheet") or self.workbook.sheetnames[0]
        date_col = params.get("date_col", "Date")
        
        sheet = self.workbook[sheet_name]
        
        # Find date column
        header_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        try:
            col_idx = header_row.index(date_col) + 1
        except ValueError:
            raise ValueError(f"Column '{date_col}' not found")
        
        # Try to parse and standardize dates
        for row_idx in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell.value:
                try:
                    # Try pandas date parsing
                    date_val = pd.to_datetime(cell.value)
                    cell.value = date_val.date()
                    cell.number_format = 'YYYY-MM-DD'
                except:
                    pass  # Keep original if parsing fails
        
        self.changes_log.append(f"Converted dates in column: {date_col}")
    
    def _add_calculated_column(self, params: Dict[str, Any]):
        """Add a new column with calculated values"""
        sheet_name = params.get("sheet") or self.workbook.sheetnames[0]
        column_name = params.get("column_name", "Calculated")
        formula_template = params.get("formula")
        
        sheet = self.workbook[sheet_name]
        
        # Add header
        last_col = sheet.max_column + 1
        sheet.cell(row=1, column=last_col, value=column_name)
        
        # Add formula to each row
        for row_idx in range(2, sheet.max_row + 1):
            formula = formula_template.replace("{ROW}", str(row_idx))
            sheet.cell(row=row_idx, column=last_col, value=formula)
        
        self.changes_log.append(f"Added calculated column: {column_name}")
    
    def save(self, output_path: str):
        """Save the modified workbook"""
        self.workbook.save(output_path)
    
    def get_diff_summary(self) -> Dict[str, Any]:
        """Get summary of changes made"""
        return {
            "changes": self.changes_log,
            "sheets": self.workbook.sheetnames,
            "total_changes": len(self.changes_log)
        }


class ActionPlanner:
    """Convert natural language requests to Excel action plans"""
    
    @staticmethod
    def parse_request(request: str) -> List[Dict[str, Any]]:
        """
        Parse natural language request into action plan
        In production, this would use OpenAI or similar
        """
        plan = []
        request_lower = request.lower()
        
        # Simple keyword matching (in production, use AI)
        if "remove duplicates" in request_lower or "duplicates" in request_lower:
            plan.append({
                "type": "remove_duplicates",
                "description": "Remove duplicate rows",
                "params": {}
            })
        
        if "trim" in request_lower or "clean" in request_lower:
            plan.append({
                "type": "trim_clean",
                "description": "Clean and trim text fields",
                "params": {"applyToAllText": True}
            })
        
        if "split" in request_lower and "name" in request_lower:
            plan.append({
                "type": "split_column",
                "description": "Split Full Name into First and Last Name",
                "params": {
                    "source_col": "Full Name",
                    "into": ["First Name", "Last Name"],
                    "delimiter": " "
                }
            })
        
        if "pivot" in request_lower:
            plan.append({
                "type": "create_pivot",
                "description": "Create pivot table summary",
                "params": {
                    "rows": ["Region"],
                    "values": [{"field": "Amount", "agg": "SUM"}],
                    "destination": "Pivot_Summary"
                }
            })
        
        if "phone" in request_lower and ("standardize" in request_lower or "format" in request_lower):
            plan.append({
                "type": "standardize_phone",
                "description": "Standardize phone number format",
                "params": {
                    "phone_col": "Phone",
                    "country_code": "234"
                }
            })
        
        if "date" in request_lower and ("convert" in request_lower or "format" in request_lower):
            plan.append({
                "type": "convert_dates",
                "description": "Convert and standardize dates",
                "params": {
                    "date_col": "Date"
                }
            })
        
        return plan

